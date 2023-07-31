# -*- coding: utf-8 -*-
# @Time ： 2023/7/31 14:22
# @Auth ： JeremyChim
# @File ：len.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim/

import openpyxl as p

def ret_row():
    f = 'report.xlsx'
    wb = p.load_workbook(f)
    ws = wb.worksheets[0]
    row = ws.max_row
    # print('工作表列数：',ws.max_column)
    # print('工作表行数：',ws.max_row)

    return row

if __name__ == '__main__':
    a = ret_row()
    print(a.__class__, a)